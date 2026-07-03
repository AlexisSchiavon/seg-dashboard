"""Generador del Excel de diagnóstico de calidad de datos SEG (READ-ONLY sobre seg.db).

Uso:
    PYTHONPATH=. .venv/bin/python scripts/diagnostico_data_quality.py [salida.xlsx]

Produce 5 hojas, una por hallazgo, cada fila = un caso accionable para Talent Agency.
NO escribe nada en la DB ni llama a integraciones externas.

Fase 9.8c — Detección de duplicados con umbral temporal (ver DUP_THRESHOLD_DAYS).
"""
import sys
import sqlite3
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB = "seg.db"
DEFAULT_OUT = "/tmp/diagnostico_data_seg_2026-07-02_v2.xlsx"
TODAY = date(2026, 7, 2)
PIPE = "https://talentagency.pipedrive.com/deal/{}"
TRELLO = "https://trello.com/c/{}"
MON = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Umbral temporal 14 días (9.8c): campañas con el mismo talent+brand+monto separadas
# por más de 2 semanas (medido sobre add_time = fecha de creación del deal) son
# campañas distintas (recurrencia mensual/bimestral), NO duplicados de importación.
# Confirmado con Alexis en Fase 9.8. Los duplicados reales del artefacto de importación
# del 8-dic-2025 tienen add_time gap = 0.
DUP_THRESHOLD_DAYS = 14

con = sqlite3.connect(DB)
con.row_factory = sqlite3.Row
cur = con.cursor()


def parse_d(s):
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except ValueError:
        return None


def fmt_d(s):
    d = parse_d(s)
    return f"{d.day:02d} {MON[d.month]} {d.year}" if d else ""


TAL = {r["id"]: r["name"] for r in cur.execute("SELECT id,name FROM talents")}


def tname(tid):
    if tid is None:
        return "Sin talento asignado"
    return TAL.get(tid, "Sin talento asignado")


HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="2563EB", underline="single")
MONEY = '$#,##0.00'


def add_sheet(wb, title, headers, rows, money_cols=(), link_cols=(), col_widths=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    for ri in range(2, len(rows) + 2):
        for ci in money_cols:
            ws.cell(ri, ci).number_format = MONEY
        for ci in link_cols:
            cell = ws.cell(ri, ci)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = LINK_FONT
    ws.freeze_panes = "A2"
    for ci in range(1, len(headers) + 1):
        letter = get_column_letter(ci)
        if col_widths and ci - 1 < len(col_widths) and col_widths[ci - 1]:
            ws.column_dimensions[letter].width = col_widths[ci - 1]
        else:
            ws.column_dimensions[letter].width = 18
    return ws


wb = Workbook()
wb.remove(wb.active)
summary = {}

# ============================================================ HOJA 1: Duplicados
won = list(cur.execute(
    "SELECT id,pipedrive_id,title,value,talent_id,won_time,add_time FROM deals WHERE status='won'"
))
groups = {}
for d in won:
    key = (d["title"].strip().lower(), d["talent_id"] if d["talent_id"] is not None else -1, d["value"])
    groups.setdefault(key, []).append(d)


def sort_key(m):
    wt = parse_d(m["won_time"])
    return (wt is None, wt or date.max, m["pipedrive_id"])


h1_rows = []
h1_total = 0.0
for key, members in groups.items():
    if len(members) < 2:
        continue
    ordered = sorted(members, key=sort_key)
    orig = ordered[0]
    orig_add = parse_d(orig["add_time"])
    for dup in ordered[1:]:
        dup_add = parse_d(dup["add_time"])
        # Umbral temporal (9.8c): >14 días de separación en add_time → campaña
        # recurrente distinta, NO duplicado. Se excluye del reporte.
        if orig_add and dup_add and abs((dup_add - orig_add).days) > DUP_THRESHOLD_DAYS:
            continue
        wo, wd = parse_d(orig["won_time"]), parse_d(dup["won_time"])
        days_between = abs((wd - wo).days) if wo and wd else ""
        # Acción específica: el artefacto de importación del 8-dic-2025 es el grueso;
        # los demás son creaciones duplicadas el mismo día pero de otra fecha.
        if wo == date(2025, 12, 8):
            accion = "Duplicado de importación del 8 dic 2025 — archivar el más reciente (mantener el original)"
        else:
            accion = "Creado el mismo día — verificar si es duplicado y archivar el más reciente"
        h1_total += dup["value"] or 0
        h1_rows.append([
            orig["pipedrive_id"], dup["pipedrive_id"], orig["title"], tname(orig["talent_id"]),
            orig["value"], fmt_d(orig["won_time"]), fmt_d(dup["won_time"]), days_between,
            PIPE.format(orig["pipedrive_id"]), PIPE.format(dup["pipedrive_id"]),
            accion,
        ])
h1_rows.sort(key=lambda r: -(r[4] or 0))
add_sheet(wb, "Duplicados Pipedrive",
    ["deal_id_original", "deal_id_duplicado", "titulo", "talent_name", "value",
     "won_time_original", "won_time_duplicado", "days_between",
     "link_pipedrive_original", "link_pipedrive_duplicado", "accion_recomendada"],
    h1_rows, money_cols=(5,), link_cols=(9, 10),
    col_widths=[16, 17, 34, 22, 14, 16, 17, 13, 42, 42, 48])
summary["1 Duplicados Pipedrive"] = (len(h1_rows), h1_total)


def deal_by_localid(lid):
    return cur.execute("SELECT * FROM deals WHERE id=?", (lid,)).fetchone()


# ============================================================ HOJA 2: Cobranza vencida
# 9.8d: SOLO la columna Trello "Cobrar" (list_state='cobranza') vencida, con el mismo
# saneo que el widget del PDF (9.8b): excluye fechas imposibles (→ Hoja 5), montos en
# 0, y >180 días (stale / probablemente ya cobrado). Contrato/Enviar factura/Firmar
# (ejecución), Enviar encuesta (post-cobro→cerrado) y Otros pendientes quedan fuera.
CUT_180 = TODAY - timedelta(days=180)
cards = list(cur.execute("SELECT * FROM trello_cards WHERE list_state='cobranza'"))
h2_rows = []
h2_total = 0.0
for c in cards:
    if c["deal_id"] is None:
        continue
    d = deal_by_localid(c["deal_id"])
    if not d or d["status"] != "won":
        continue
    cd = parse_d(c["collection_date"])
    if not cd or cd >= TODAY:
        continue
    ad = parse_d(d["add_time"])
    if ad and cd < ad:
        continue  # fecha imposible → se reporta en Hoja 5
    if (d["value"] or 0) <= 0:
        continue  # nada real que cobrar
    if cd < CUT_180:
        continue  # >180 días — stale / probablemente ya cobrado
    dias = (TODAY - cd).days
    if dias >= 180:
        acc = "Verificar si ya se cobró y no se actualizó. Marcar cerrado o reprogramar."
    elif dias >= 91:
        acc = "Escalación con marca. Actualizar fecha o marcar cerrado."
    elif dias >= 31:
        acc = "Contactar marca. Actualizar collection_date."
    else:
        acc = "En seguimiento normal. Confirmar fecha realista."
    h2_total += d["value"] or 0
    h2_rows.append([
        c["id"], c["trello_card_id"], d["pipedrive_id"], d["title"], tname(d["talent_id"]),
        d["value"], c["list_state"], fmt_d(c["collection_date"]), dias,
        TRELLO.format(c["trello_card_id"]), PIPE.format(d["pipedrive_id"]), acc,
    ])
h2_rows.sort(key=lambda r: -r[8])
add_sheet(wb, "Cobranza vencida Trello",
    ["card_id", "trello_card_id", "deal_id", "titulo_deal", "talent_name", "value",
     "list_state", "collection_date", "dias_vencido", "link_trello", "link_pipedrive",
     "accion_recomendada"],
    h2_rows, money_cols=(6,), link_cols=(10, 11),
    col_widths=[8, 26, 12, 34, 22, 14, 12, 16, 12, 46, 42, 56])
summary["2 Cobranza vencida Trello"] = (len(h2_rows), h2_total)

# ============================================================ HOJA 3: Ganados en $0
z = list(cur.execute(
    "SELECT * FROM deals WHERE status='won' AND (value=0 OR value IS NULL) ORDER BY won_time"
))
h3_rows = []
for d in z:
    has_card = cur.execute(
        "SELECT 1 FROM trello_cards WHERE deal_id=? LIMIT 1", (d["id"],)).fetchone() is not None
    h3_rows.append([
        d["pipedrive_id"], d["title"], tname(d["talent_id"]), fmt_d(d["won_time"]),
        "Sí" if has_card else "No", PIPE.format(d["pipedrive_id"]),
        "Agregar el monto real de la campaña en Pipedrive",
    ])
add_sheet(wb, "Deals ganados en 0",
    ["deal_id", "titulo", "talent_name", "won_time", "tiene_card_trello",
     "link_pipedrive", "accion_recomendada"],
    h3_rows, link_cols=(6,),
    col_widths=[12, 40, 24, 16, 16, 42, 48])
summary["3 Deals ganados en $0"] = (len(h3_rows), 0.0)

# ============================================================ HOJA 4: Ganados sin Trello
h4 = list(cur.execute(
    "SELECT * FROM deals d WHERE d.status='won' "
    "AND NOT EXISTS (SELECT 1 FROM trello_cards t WHERE t.deal_id=d.id) ORDER BY d.value DESC"
))
h4_rows = []
h4_total = 0.0
for d in h4:
    h4_total += d["value"] or 0
    h4_rows.append([
        d["pipedrive_id"], d["title"], tname(d["talent_id"]), d["value"],
        fmt_d(d["won_time"]), PIPE.format(d["pipedrive_id"]),
        "Crear tarjeta en Trello Board Admin TA con collection_date realista",
    ])
add_sheet(wb, "Deals ganados sin Trello",
    ["deal_id", "titulo", "talent_name", "value", "won_time", "link_pipedrive",
     "accion_recomendada"],
    h4_rows, money_cols=(4,), link_cols=(6,),
    col_widths=[12, 40, 24, 14, 16, 42, 56])
summary["4 Deals ganados sin Trello"] = (len(h4_rows), h4_total)

# ============================================================ HOJA 5: Fechas imposibles
allcards = list(cur.execute("SELECT * FROM trello_cards"))
h5_rows = []
h5_total = 0.0
for c in allcards:
    if c["deal_id"] is None:
        continue
    d = deal_by_localid(c["deal_id"])
    if not d:
        continue
    cd = parse_d(c["collection_date"])
    ad = parse_d(d["add_time"])
    if not cd or not ad or cd >= ad:
        continue
    diff = (ad - cd).days
    h5_total += d["value"] or 0
    h5_rows.append([
        c["id"], c["trello_card_id"], d["pipedrive_id"], d["title"], tname(d["talent_id"]),
        d["value"], fmt_d(d["add_time"]), fmt_d(c["collection_date"]), diff,
        TRELLO.format(c["trello_card_id"]),
        "Corregir collection_date en Trello - la fecha actual es anterior a la creación del deal",
    ])
h5_rows.sort(key=lambda r: -r[8])
add_sheet(wb, "Fechas imposibles Trello",
    ["card_id", "trello_card_id", "deal_id", "titulo_deal", "talent_name", "value",
     "add_time", "collection_date", "diferencia_dias", "link_trello", "accion_recomendada"],
    h5_rows, money_cols=(6,), link_cols=(10,),
    col_widths=[8, 26, 12, 34, 22, 14, 16, 16, 14, 46, 62])
summary["5 Fechas imposibles Trello"] = (len(h5_rows), h5_total)

out = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
wb.save(out)
con.close()

print("SAVED:", out)
print()
for k, (n, tot) in summary.items():
    print(f"{k:34s} filas={n:4d}  $afectado={tot:>15,.2f}")
